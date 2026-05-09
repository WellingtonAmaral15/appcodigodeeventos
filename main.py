import os
import sys
import unicodedata

os.environ.setdefault("KIVY_NO_FILELOG", "1")
from functools import partial

from kivy.app import App
from kivy.graphics import Color, Rectangle
from kivy.lang import Builder
from kivy.uix.button import ButtonBehavior
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image
from kivy.uix.label import Label
from kivy.uix.screenmanager import Screen
from openpyxl import load_workbook

APP_DIR = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
NUMERIC_SYSTEMS = {"EIMS NUM.xlsm", "TMS NUM.xlsm", "FREIO KNORR NUM.xlsm"}
CODE_KEY = "CODIGO"
DESCRIPTION_KEY = "DESCRICAO"
COMPONENT_KEY = "COMPONENTE"


def resource_path(*parts):
    return os.path.join(APP_DIR, *parts)


def normalize_code(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    return str(value).strip()


def normalize_key(value):
    text = normalize_code(value).upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text


def read_event_rows(filename):
    workbook_path = resource_path("arquivos", filename)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active

    headers = [normalize_key(cell.value) for cell in sheet[1]]
    rows = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, values))
        code = normalize_code(item.get(CODE_KEY))
        if not code:
            continue

        rows.append(
            {
                "codigo": code,
                "descricao": normalize_code(item.get(DESCRIPTION_KEY)),
                "componente": normalize_code(item.get(COMPONENT_KEY)),
            }
        )

    workbook.close()
    return rows


def format_visible_code(code, filename):
    if filename == "FREIO KNORR NUM.xlsm" and code.isdigit():
        return code.zfill(4)
    return code


def find_event(rows, code, filename):
    code = normalize_code(code)
    if filename in NUMERIC_SYSTEMS:
        code = code.lstrip("0") or "0"

    for row in rows:
        row_code = row["codigo"]
        compare_code = row_code.lstrip("0") or "0" if filename in NUMERIC_SYSTEMS else row_code
        if compare_code == code:
            return row

    return None


class SistemaPage(Screen):
    pass


class SelecionarPage(Screen):
    pass


class ImageButton(ButtonBehavior, Image):
    pass


class LabelButton(ButtonBehavior, Label):
    pass


class BannerSistema(GridLayout):
    def __init__(self, **kwargs):
        super().__init__()
        self.cols = 3

        imagem = kwargs["imagem"]
        label = kwargs["label"]
        meu_aplicativo = App.get_running_app()

        banner = FloatLayout()
        banner_imagem = ImageButton(
            pos_hint={"right": 1, "top": 0.85},
            size_hint=(1, 0.65),
            source=resource_path("icones", imagem),
            on_release=partial(meu_aplicativo.selecionar_sistema, imagem),
        )
        banner_label = LabelButton(
            text=label,
            pos_hint={"right": 1, "top": 0.15},
            size_hint=(1, 0.15),
            on_release=partial(meu_aplicativo.selecionar_sistema, imagem),
        )

        banner.add_widget(banner_imagem)
        banner.add_widget(banner_label)
        self.add_widget(banner)


class BannerSelecionar(GridLayout):
    def __init__(self, **kwargs):
        super().__init__()
        self.rows = 2

        with self.canvas:
            Color(rgb=(0, 0, 0, 1))
            self.rec = Rectangle(size=self.size, pos=self.pos)
        self.bind(pos=self.atualizar_rec, size=self.atualizar_rec)

        codigo = kwargs["codigo"]
        meu_aplicativo = App.get_running_app()

        banner = FloatLayout()
        banner_label = LabelButton(
            text=codigo,
            pos_hint={"right": 0.6, "top": 0.15},
            size_hint=(0.3, 0.15),
            on_release=partial(meu_aplicativo.preencher_codigo, codigo),
        )

        banner.add_widget(banner_label)
        self.add_widget(banner)

    def atualizar_rec(self, *args):
        self.rec.pos = self.pos
        self.rec.size = self.size


GUI = Builder.load_file(resource_path("main.kv"))


class MainApp(App):
    def build(self):
        return GUI

    def on_start(self):
        lista_icones = self.root.ids["sistemapage"].ids["scroolview_sistema"]
        for foto_icones in sorted(os.listdir(resource_path("icones"))):
            if not foto_icones.lower().endswith(".png"):
                continue

            label = foto_icones.replace(".png", "")
            banner = BannerSistema(imagem=foto_icones, label=label)
            lista_icones.add_widget(banner)

    def buscar_sistema(self, *args):
        pagina_sistema = self.root.ids["sistemapage"]
        pagina_selecionar = self.root.ids["selecionarpage"]
        sistema = pagina_sistema.ids["sistema_selecionado"].text

        if not sistema or sistema in {"Selecione um Sistema", "SELECIONE UM SISTEMA!!!"}:
            pagina_selecionar.ids["imagem_selecionar"].source = resource_path("icones2", "images.png")
            pagina_sistema.ids["sistema_selecionado"].text = "SELECIONE UM SISTEMA!!!"
            pagina_selecionar.ids["label_sistema_selecionado"].text = "SELECIONE UM SISTEMA!!!"
            return

        arquivo = f"{sistema}.xlsm"
        pagina_selecionar.ids["imagem_selecionar"].source = resource_path("icones", f"{sistema}.png")
        lista_cod_selecionar = pagina_selecionar.ids["scroolview_selecionar"]
        lista_cod_selecionar.clear_widgets()

        try:
            rows = read_event_rows(arquivo)
        except Exception as exc:
            pagina_selecionar.ids["componente"].text = "[color=#FF0000]Erro ao abrir arquivo[/color]"
            pagina_selecionar.ids["descricao"].text = f"[color=#FF0000]{exc}[/color]"
            return

        for row in rows:
            codigo = format_visible_code(row["codigo"], arquivo)
            lista_cod_selecionar.add_widget(BannerSelecionar(codigo=codigo))

        texto_input = pagina_selecionar.ids["codigo_input"]
        if arquivo in NUMERIC_SYSTEMS:
            texto_input.hint_text = "Digite o codigo numerico e clique no botao verde ao lado"
        else:
            texto_input.hint_text = "Digite o codigo alfabetico e clique no botao verde ao lado"

    def preencher_codigo(self, codigo, *args):
        self.root.ids["selecionarpage"].ids["codigo_escolhido"].text = codigo

    def selecionar_input(self, text):
        pagina_selecionar = self.root.ids["selecionarpage"]
        pagina_selecionar.ids["codigo_escolhido"].text = "Codigo"
        pagina_selecionar.ids["componente"].text = "Componente"
        pagina_selecionar.ids["descricao"].text = "Descricao da Falha"

        if text:
            pagina_selecionar.ids["codigo_escolhido"].text = text
        else:
            pagina_selecionar.ids["componente"].text = "[color=#FF0000]Codigo Invalido[/color]"
            pagina_selecionar.ids["descricao"].text = "[color=#FF0000]Digite um Codigo Valido[/color]"

    def buscar_dados(self, *args):
        pagina_selecionar = self.root.ids["selecionarpage"]
        pagina_sistema = self.root.ids["sistemapage"]
        sistema = pagina_sistema.ids["sistema_selecionado"].text
        arquivo = f"{sistema}.xlsm"
        codigo = pagina_selecionar.ids["codigo_escolhido"].text
        label_componente = pagina_selecionar.ids["componente"]
        label_descricao = pagina_selecionar.ids["descricao"]

        if not codigo or codigo == "Codigo":
            label_componente.text = ""
            label_descricao.text = "[color=#000000]Selecione ou digite um Codigo[/color]"
            return

        try:
            rows = read_event_rows(arquivo)
            event = find_event(rows, codigo, arquivo)
        except Exception as exc:
            label_componente.text = "[color=#FF0000]Erro ao buscar codigo[/color]"
            label_descricao.text = f"[color=#FF0000]{exc}[/color]"
            return

        if event is None:
            label_componente.text = "[color=#000000]Codigo Invalido[/color]"
            label_descricao.text = "[color=#000000]Codigo nao contemplado no sistema[/color]"
            return

        label_componente.text = f"[color=#000000]Componente:[/color] [b]{event['componente']}[/b]"
        label_descricao.text = f"[color=#000000]Descricao:[/color] [b]{event['descricao']}[/b]"

    def mudar_tela(self, id_tela):
        self.root.ids["screen_manager"].current = id_tela

    def selecionar_sistema(self, imagem, *args):
        pagina_selecionar = self.root.ids["selecionarpage"]
        pagina_selecionar.ids["codigo_escolhido"].text = "Codigo"
        pagina_selecionar.ids["componente"].text = "Componente"
        pagina_selecionar.ids["descricao"].text = "Descricao da Falha"

        sistema = imagem.replace(".png", "")
        self.root.ids["sistemapage"].ids["sistema_selecionado"].text = sistema
        pagina_selecionar.ids["label_sistema_selecionado"].text = sistema

    def fechar_aplicativo(self):
        self.stop()


if __name__ == "__main__":
    MainApp().run()
