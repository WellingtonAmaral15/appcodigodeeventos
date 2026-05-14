import argparse
import json
import mimetypes
import os
import socket
import unicodedata
from functools import lru_cache
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parent
ARQUIVOS_DIR = APP_DIR / "arquivos"
ICONES_DIR = APP_DIR / "icones"
ICONES2_DIR = APP_DIR / "icones2"

NUMERIC_SYSTEMS = {"EIMS NUM 2005.xlsm", "TMS NUM 3000.xlsm", "FREIO KNORR NUM 3000 4000 E 5000.xlsm"}
CODE_KEY = "CODIGO"
DESCRIPTION_KEY = "DESCRICAO"
COMPONENT_KEY = "COMPONENTE"
SYSTEM_DISPLAY_ORDER = [
    "APU 2005 E 3000",
    "CVS 4000 E 5000",
    "EIMS 2005",
    "EIMS NUM 2005",
    "FREIO KNORR 3000 4000 E 5000",
    "FREIO KNORR NUM 3000 4000 E 5000",
    "VVVF 2005 E 3000",
    "INVERSOR DE TRACAO 4000 E 5000",
    "TMS 3000",
    "TMS NUM 3000",
    "EVR 4000 E 5000",
    "PERFORMANCE 3000",
]


HTML_PAGE = r"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Codigo de Eventos</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #eef3f5;
      --panel: #ffffff;
      --ink: #11242d;
      --muted: #61727b;
      --line: #d6e0e4;
      --accent: #006d7e;
      --accent-strong: #004f5c;
      --warn: #b3261e;
      --ok: #1f7a48;
      --soft: #f7fafb;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      background: var(--bg);
      color: var(--ink);
      font-family: Segoe UI, system-ui, -apple-system, BlinkMacSystemFont, sans-serif;
    }
    header {
      position: sticky;
      top: 0;
      z-index: 5;
      background: #001824;
      color: #fff;
      border-bottom: 3px solid var(--accent);
    }
    .topbar {
      display: flex;
      align-items: center;
      gap: 12px;
      width: min(1180px, 100%);
      margin: 0 auto;
      padding: 10px 14px;
    }
    .brand-icon {
      width: 42px;
      height: 42px;
      object-fit: contain;
      flex: 0 0 auto;
    }
    h1 {
      margin: 0;
      font-size: 1.08rem;
      line-height: 1.2;
      letter-spacing: 0;
    }
    main {
      width: min(1180px, 100%);
      margin: 0 auto;
      padding: 14px;
      display: grid;
      grid-template-columns: minmax(260px, 380px) minmax(0, 1fr);
      gap: 14px;
    }
    section {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      min-width: 0;
    }
    .systems {
      padding: 12px;
      display: grid;
      gap: 10px;
      align-content: start;
    }
    .panel-title {
      margin: 0 0 4px;
      font-size: .9rem;
      color: var(--muted);
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: .02em;
    }
    .system-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
    }
    .system-button {
      appearance: none;
      display: grid;
      grid-template-rows: 58px auto;
      gap: 6px;
      align-items: center;
      justify-items: center;
      min-height: 102px;
      padding: 9px 7px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--soft);
      color: var(--ink);
      cursor: pointer;
      font: inherit;
      text-align: center;
    }
    .system-button:hover,
    .system-button:focus-visible {
      outline: 2px solid rgba(0, 109, 126, .22);
      border-color: var(--accent);
      background: #fff;
    }
    .system-button.active {
      border-color: var(--accent);
      background: #e8f5f6;
      box-shadow: inset 0 0 0 1px var(--accent);
    }
    .system-button img {
      width: 100%;
      max-width: 92px;
      height: 58px;
      object-fit: contain;
    }
    .system-button span {
      width: 100%;
      min-height: 28px;
      display: flex;
      align-items: center;
      justify-content: center;
      overflow-wrap: anywhere;
      font-size: .82rem;
      font-weight: 700;
      line-height: 1.15;
    }
    .workspace {
      padding: 14px;
      display: grid;
      gap: 12px;
      align-content: start;
    }
    .selected-system {
      display: grid;
      grid-template-columns: 92px minmax(0, 1fr);
      gap: 12px;
      align-items: center;
      min-height: 98px;
      padding: 10px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--soft);
    }
    .selected-system img {
      width: 92px;
      height: 72px;
      object-fit: contain;
    }
    .selected-name {
      margin: 0;
      font-size: 1.2rem;
      line-height: 1.2;
      overflow-wrap: anywhere;
    }
    .hint {
      margin: 4px 0 0;
      color: var(--muted);
      font-size: .92rem;
    }
    .search-row {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 8px;
    }
    input[type="search"] {
      width: 100%;
      min-height: 44px;
      border: 1px solid var(--line);
      border-radius: 7px;
      padding: 8px 10px;
      color: var(--ink);
      background: #fff;
      font: inherit;
    }
    button.primary {
      min-height: 44px;
      border: 0;
      border-radius: 7px;
      padding: 0 18px;
      background: var(--accent);
      color: #fff;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      white-space: nowrap;
    }
    button.primary:hover,
    button.primary:focus-visible {
      background: var(--accent-strong);
      outline: 2px solid rgba(0, 109, 126, .22);
    }
    .codes-wrap { display: grid; gap: 8px; }
    .codes-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      color: var(--muted);
      font-size: .9rem;
      font-weight: 700;
    }
    .codes-list {
      display: flex;
      gap: 8px;
      overflow-x: auto;
      padding-bottom: 4px;
      min-height: 46px;
      scrollbar-width: thin;
    }
    .code-chip {
      appearance: none;
      flex: 0 0 auto;
      min-width: 86px;
      min-height: 38px;
      border: 1px solid var(--line);
      border-radius: 7px;
      padding: 6px 10px;
      background: #fff;
      color: var(--ink);
      font: inherit;
      font-weight: 700;
      cursor: pointer;
    }
    .code-chip:hover,
    .code-chip:focus-visible {
      border-color: var(--accent);
      outline: 2px solid rgba(0, 109, 126, .18);
    }
    .result {
      display: grid;
      gap: 10px;
      min-height: 150px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
      background: #fff;
    }
    .result-line { display: grid; gap: 4px; }
    .result-line strong {
      color: var(--muted);
      font-size: .86rem;
      text-transform: uppercase;
      letter-spacing: .02em;
    }
    .result-line span {
      font-size: 1.05rem;
      line-height: 1.35;
      overflow-wrap: anywhere;
    }
    .status {
      min-height: 22px;
      color: var(--muted);
      font-size: .92rem;
    }
    .status.error { color: var(--warn); font-weight: 700; }
    .status.ok { color: var(--ok); font-weight: 700; }
    .empty { color: var(--muted); font-weight: 600; padding: 8px 0; }
    @media (max-width: 760px) {
      main { grid-template-columns: 1fr; padding: 10px; }
      .system-grid { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .selected-system { grid-template-columns: 74px minmax(0, 1fr); }
      .selected-system img { width: 74px; height: 58px; }
      .search-row { grid-template-columns: 1fr; }
      button.primary { width: 100%; }
    }
    @media (max-width: 460px) {
      h1 { font-size: .98rem; }
      .system-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .system-button { min-height: 96px; }
    }
  </style>
</head>
<body>
  <header>
    <div class="topbar">
      <img class="brand-icon" src="/static/icones2/ferramenta.png" alt="">
      <h1>Ferramenta de Analise de Falhas</h1>
    </div>
  </header>

  <main>
    <section class="systems" aria-labelledby="systems-title">
      <h2 id="systems-title" class="panel-title">Sistemas</h2>
      <div id="systems" class="system-grid"></div>
    </section>

    <section class="workspace" aria-live="polite">
      <div class="selected-system">
        <img id="selected-image" src="/static/icones2/images.png" alt="">
        <div>
          <h2 id="selected-name" class="selected-name">Selecione um Sistema</h2>
          <p id="selected-hint" class="hint">Escolha um sistema para carregar os codigos.</p>
        </div>
      </div>

      <form id="search-form" class="search-row">
        <input id="code-input" type="search" autocomplete="off" placeholder="Digite o codigo">
        <button class="primary" type="submit">Buscar</button>
      </form>

      <div class="codes-wrap">
        <div class="codes-head">
          <span>Codigos</span>
          <span id="codes-count">0</span>
        </div>
        <div id="codes-list" class="codes-list">
          <span class="empty">Nenhum sistema selecionado.</span>
        </div>
      </div>

      <div id="status" class="status"></div>

      <div class="result">
        <div class="result-line">
          <strong>Codigo</strong>
          <span id="result-code">-</span>
        </div>
        <div class="result-line">
          <strong>Componente</strong>
          <span id="result-component">-</span>
        </div>
        <div class="result-line">
          <strong>Descricao da Falha</strong>
          <span id="result-description">-</span>
        </div>
      </div>
    </section>
  </main>

  <script>
    const state = { systems: [], selected: null, codes: [] };
    const el = {
      systems: document.getElementById("systems"),
      selectedImage: document.getElementById("selected-image"),
      selectedName: document.getElementById("selected-name"),
      selectedHint: document.getElementById("selected-hint"),
      codeInput: document.getElementById("code-input"),
      searchForm: document.getElementById("search-form"),
      codesList: document.getElementById("codes-list"),
      codesCount: document.getElementById("codes-count"),
      status: document.getElementById("status"),
      resultCode: document.getElementById("result-code"),
      resultComponent: document.getElementById("result-component"),
      resultDescription: document.getElementById("result-description")
    };

    function asset(path) {
      return path.split("/").map(encodeURIComponent).join("/");
    }
    function setStatus(message, kind = "") {
      el.status.textContent = message;
      el.status.className = `status ${kind}`.trim();
    }
    function clearResult() {
      el.resultCode.textContent = "-";
      el.resultComponent.textContent = "-";
      el.resultDescription.textContent = "-";
    }
    async function fetchJson(url) {
      const response = await fetch(url);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "Falha ao carregar dados.");
      return data;
    }
    async function loadSystems() {
      setStatus("Carregando sistemas...");
      const data = await fetchJson("/api/systems");
      state.systems = data.systems;
      renderSystems();
      setStatus(`${state.systems.length} sistemas carregados.`, "ok");
    }
    function renderSystems() {
      el.systems.innerHTML = "";
      for (const system of state.systems) {
        const button = document.createElement("button");
        button.type = "button";
        button.className = "system-button";
        button.dataset.system = system.name;
        button.innerHTML = `<img src="/static/icones/${asset(system.image)}" alt=""><span>${system.name}</span>`;
        button.addEventListener("click", () => selectSystem(system.name));
        el.systems.appendChild(button);
      }
    }
    async function selectSystem(systemName) {
      const system = state.systems.find(item => item.name === systemName);
      if (!system) return;
      state.selected = system;
      state.codes = [];
      clearResult();
      el.codeInput.value = "";
      el.selectedName.textContent = system.name;
      el.selectedImage.src = `/static/icones/${asset(system.image)}`;
      el.selectedHint.textContent = system.numeric ? "Digite ou toque em um codigo numerico." : "Digite ou toque em um codigo alfabetico.";
      document.querySelectorAll(".system-button").forEach(button => {
        button.classList.toggle("active", button.dataset.system === system.name);
      });
      el.codesList.innerHTML = `<span class="empty">Carregando codigos...</span>`;
      el.codesCount.textContent = "0";
      setStatus("Carregando codigos...");
      try {
        const data = await fetchJson(`/api/codes?system=${encodeURIComponent(system.name)}`);
        state.codes = data.codes;
        renderCodes();
        setStatus(`${system.name} selecionado.`, "ok");
      } catch (error) {
        el.codesList.innerHTML = `<span class="empty">Nao foi possivel carregar os codigos.</span>`;
        setStatus(error.message, "error");
      }
    }
    function renderCodes() {
      el.codesCount.textContent = String(state.codes.length);
      el.codesList.innerHTML = "";
      if (!state.codes.length) {
        el.codesList.innerHTML = `<span class="empty">Nenhum codigo encontrado.</span>`;
        return;
      }
      for (const code of state.codes) {
        const button = document.createElement("button");
        button.type = "button";
        button.className = "code-chip";
        button.textContent = code;
        button.addEventListener("click", () => {
          el.codeInput.value = code;
          searchCode();
        });
        el.codesList.appendChild(button);
      }
    }
    async function searchCode() {
      if (!state.selected) {
        setStatus("Selecione um sistema antes de buscar.", "error");
        return;
      }
      const code = el.codeInput.value.trim();
      if (!code) {
        clearResult();
        setStatus("Digite ou selecione um codigo.", "error");
        return;
      }
      setStatus("Buscando codigo...");
      try {
        const data = await fetchJson(`/api/search?system=${encodeURIComponent(state.selected.name)}&code=${encodeURIComponent(code)}`);
        el.resultCode.textContent = data.codigo || code;
        el.resultComponent.textContent = data.componente || "-";
        el.resultDescription.textContent = data.descricao || "-";
        setStatus("Codigo encontrado.", "ok");
      } catch (error) {
        el.resultCode.textContent = code;
        el.resultComponent.textContent = "Codigo invalido";
        el.resultDescription.textContent = "Codigo nao contemplado no sistema";
        setStatus(error.message, "error");
      }
    }
    el.searchForm.addEventListener("submit", event => {
      event.preventDefault();
      searchCode();
    });
    loadSystems().catch(error => setStatus(error.message, "error"));
  </script>
</body>
</html>
"""


def normalize_code(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def normalize_key(value):
    text = normalize_code(value).upper()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def format_visible_code(code, filename):
    if filename == "FREIO KNORR NUM 3000 4000 E 5000.xlsm" and code.isdigit():
        return code.zfill(4)
    return code


def get_system_filename(system_name):
    filename = f"{system_name}.xlsm"
    path = ARQUIVOS_DIR / filename
    if not path.is_file():
        raise ValueError("Sistema nao encontrado.")
    return filename


@lru_cache(maxsize=64)
def read_event_rows(filename):
    workbook_path = ARQUIVOS_DIR / filename
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [normalize_key(cell.value) for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, values))
        code = normalize_code(item.get(CODE_KEY))
        if not code:
            continue
        rows.append({
            "codigo": code,
            "codigo_visivel": format_visible_code(code, filename),
            "descricao": normalize_code(item.get(DESCRIPTION_KEY)),
            "componente": normalize_code(item.get(COMPONENT_KEY)),
        })
    workbook.close()
    return rows


def find_event(rows, code, filename):
    code = normalize_code(code)
    if filename in NUMERIC_SYSTEMS:
        code = code.lstrip("0") or "0"
    else:
        code = normalize_key(code)
    for row in rows:
        row_code = row["codigo"]
        if filename in NUMERIC_SYSTEMS:
            compare_code = row_code.lstrip("0") or "0"
        else:
            compare_code = normalize_key(row_code)
        if compare_code == code:
            return row
    return None


def list_systems():
    systems = []
    order_map = {normalize_key(name): index for index, name in enumerate(SYSTEM_DISPLAY_ORDER)}
    icon_paths = sorted(
        ICONES_DIR.glob("*.png"),
        key=lambda path: (order_map.get(normalize_key(path.stem), len(order_map)), path.name.casefold()),
    )
    for icon_path in icon_paths:
        name = icon_path.stem
        filename = f"{name}.xlsm"
        if not (ARQUIVOS_DIR / filename).is_file():
            continue
        systems.append({
            "name": name,
            "image": icon_path.name,
            "filename": filename,
            "numeric": filename in NUMERIC_SYSTEMS,
        })
    return systems


def json_bytes(payload):
    return json.dumps(payload, ensure_ascii=False).encode("utf-8")


def safe_static_path(base_dir, relative_path):
    decoded = unquote(relative_path).replace("\\", "/").lstrip("/")
    path = (base_dir / decoded).resolve()
    base = base_dir.resolve()
    try:
        path.relative_to(base)
    except ValueError:
        return None
    if not path.is_file():
        return None
    return path


def get_lan_ip():
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            return sock.getsockname()[0]
    except OSError:
        return "127.0.0.1"


class WebAppHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)
        try:
            if path == "/":
                self.send_html(HTML_PAGE)
            elif path == "/api/systems":
                self.send_json({"systems": list_systems()})
            elif path == "/api/codes":
                self.handle_codes(query)
            elif path == "/api/search":
                self.handle_search(query)
            elif path.startswith("/static/icones/"):
                self.send_static(ICONES_DIR, path.removeprefix("/static/icones/"))
            elif path.startswith("/static/icones2/"):
                self.send_static(ICONES2_DIR, path.removeprefix("/static/icones2/"))
            else:
                self.send_error_json(HTTPStatus.NOT_FOUND, "Rota nao encontrada.")
        except ValueError as exc:
            self.send_error_json(HTTPStatus.BAD_REQUEST, str(exc))
        except Exception as exc:
            self.send_error_json(HTTPStatus.INTERNAL_SERVER_ERROR, f"Erro interno: {exc}")

    def handle_codes(self, query):
        system = query.get("system", [""])[0]
        if not system:
            raise ValueError("Informe o sistema.")
        filename = get_system_filename(system)
        rows = read_event_rows(filename)
        self.send_json({"system": system, "codes": [row["codigo_visivel"] for row in rows]})

    def handle_search(self, query):
        system = query.get("system", [""])[0]
        code = query.get("code", [""])[0]
        if not system:
            raise ValueError("Informe o sistema.")
        if not code:
            raise ValueError("Informe o codigo.")
        filename = get_system_filename(system)
        rows = read_event_rows(filename)
        event = find_event(rows, code, filename)
        if event is None:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Codigo nao encontrado.")
            return
        self.send_json({
            "codigo": format_visible_code(event["codigo"], filename),
            "componente": event["componente"],
            "descricao": event["descricao"],
        })

    def send_html(self, html):
        data = html.encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_json(self, payload, status=HTTPStatus.OK):
        data = json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_error_json(self, status, message):
        self.send_json({"error": message}, status)

    def send_static(self, base_dir, relative_path):
        path = safe_static_path(base_dir, relative_path)
        if path is None:
            self.send_error_json(HTTPStatus.NOT_FOUND, "Arquivo nao encontrado.")
            return
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        data = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "public, max-age=3600")
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format, *args):
        return


def main():
    parser = argparse.ArgumentParser(description="Codigo de Eventos para acesso web.")
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", default=int(os.environ.get("PORT", "8000")), type=int)
    args = parser.parse_args()
    server = ThreadingHTTPServer((args.host, args.port), WebAppHandler)
    lan_ip = get_lan_ip()
    print("Codigo de Eventos Web em execucao.")
    print(f"No computador: http://127.0.0.1:{args.port}")
    print(f"No celular:    http://{lan_ip}:{args.port}")
    print("Use Ctrl+C para parar.")
    server.serve_forever()


if __name__ == "__main__":
    main()
